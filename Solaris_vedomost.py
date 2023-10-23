# -*- coding: utf-8 -*-

#
# Created by: PyQt5 UI code generator 5.10.1
#
# WARNING! All changes made in this file will be lost!

import re
from PyQt5.QtWidgets import (QMainWindow, QMessageBox, QFileDialog, QLineEdit, QRadioButton, QInputDialog)
import rashet_sekcii
import to_exl_main_mat
from PyQt5 import QtCore, QtWidgets
import logg_solaris
import sys  # sys нужен для передачи argv в QApplication
import xlrd
import materials_simple
import tools
# import logging
import openpyxl
from collections import defaultdict
import sys
import os
import shutil
from PyQt5.QtCore import QDateTime
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QLabel, QVBoxLayout, QWidget, QMessageBox

ORGANIZATION_NAME = 'PIK Solaris'
ORGANIZATION_DOMAIN = 'www.piksolaris.ru'
APPLICATION_NAME = 'Solaris_vedomost'
SETTINGS_TRAY = 'settings/tray'


class ExampleApp(QMainWindow):
    check_box = None

    def __init__(self, parent=None):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        print('Начало программы')
        super(ExampleApp, self).__init__(parent)
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(500, 300)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(20, 60, 100, 1000))
        self.widget.setObjectName("widget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.widget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        # Открываем файл
        self.pushButton2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton2.setGeometry(QtCore.QRect(120, 100, 100, 20))
        self.pushButton2.setObjectName("pushButton2")
        self.pushButton2.clicked.connect(self.showDialog)  # если кнопка нажата

        # выпадающий список
        self.zadacha = QtWidgets.QComboBox(self.centralwidget)
        self.zadacha.setObjectName("material")
        self.zadacha.setGeometry(QtCore.QRect(240, 100, 200, 20))
        self.zadacha.addItem("Создать ЦВ")
        self.zadacha.addItem("Переработать ЦВ")
        self.zadacha.addItem("Объеденить")

        # Начинаем расчет (Запуск, Начать)
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(20, 165, 75, 20))
        self.pushButton.setStyleSheet('background: rgb(131, 230, 144);')
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.nachalo)  # если кнопка нажата

        # кнопка расчета материалов
        self.pushButton3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton3.setGeometry(QtCore.QRect(220, 165, 150, 20))
        self.pushButton3.setObjectName("pushButton3")
        self.pushButton3.clicked.connect(self.handleButton)  # если кнопка нажата

        # кнопка расчета материалов
        self.pushButton4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton4.setGeometry(QtCore.QRect(220, 185, 170, 20))
        self.pushButton4.setObjectName("pushButton4")
        self.pushButton4.clicked.connect(self.handleButton2)  # если кнопка нажата

        self.label_zakaz = QtWidgets.QLabel(self.centralwidget)
        self.label_zakaz.setGeometry(QtCore.QRect(20, 20, 100, 10))
        self.label_zakaz.setObjectName("label")
        self.label_zakaz.setText('№ Заказа')
        self.line_zakaz = QLineEdit(self)
        self.line_zakaz.move(20, 40)
        self.line_zakaz.resize(80, 20)

        self.label_project = QtWidgets.QLabel(self.centralwidget)
        self.label_project.setGeometry(QtCore.QRect(130, 20, 100, 10))
        self.label_project.setObjectName("label")
        self.label_project.setText('№ Проекта')
        self.line_project = QLineEdit(self)
        self.line_project.move(130, 40)
        self.line_project.resize(80, 20)

        self.label_project_name = QtWidgets.QLabel(self.centralwidget)
        self.label_project_name.setGeometry(QtCore.QRect(250, 20, 160, 10))
        self.label_project_name.setObjectName("label")
        self.label_project_name.setText('№ Наименование проекта')
        self.line_project_name = QLineEdit(self)
        self.line_project_name.move(250, 40)
        self.line_project_name.resize(180, 20)

        self.label_stage = QtWidgets.QLabel(self.centralwidget)
        self.label_stage.setGeometry(QtCore.QRect(20, 60, 80, 40))
        self.label_stage.setObjectName("label")
        self.label_stage.setText('№ Этапа')
        self.line_stage = QLineEdit(self)
        self.line_stage.move(20, 100)
        self.line_stage.resize(80, 20)

        self.v1button = QRadioButton("Питон Кама", self)
        self.v1button.setGeometry(QtCore.QRect(20, 200, 100, 10))
        self.v2button = QRadioButton("Солярис", self)
        self.v2button.setGeometry(QtCore.QRect(20, 230, 100, 10))
        self.v2button.setChecked(True)

        MainWindow.setCentralWidget(self.centralwidget)

        self.fname = ''
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.atribut = []  # Номер заказа, номер проекта

    # Расчет материалов
    def handleButton(self):
        next = ExampleApp2(self)
        next.show()

    # Разница в спецификациях
    def handleButton2(self):
        next = ExcelProcessingApp(self)
        next.show()


    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Solaris specification  V-231023-01"))
        self.pushButton.setText(_translate("MainWindow", "Пуск"))
        self.pushButton2.setText(_translate("MainWindow", "Открыть файл"))
        self.pushButton3.setText(_translate("MainWindow", "Расчет материалов"))
        self.pushButton4.setText(_translate("MainWindow", "Разница в спецификациях"))

    def showDialog(self):
        self.fname = QFileDialog.getOpenFileName(self, 'Открыть файл', '', "Exel (*.xlsx *.xls)", )[0]

        if self.fname == '':  # ничего не выбрано показывает предупреждение
            print('файл не выбран')
            message = 'Вы не выбрали файл. Хотите выбрать? \n Или хотите закончить расчет? \n'
            reply = QtWidgets.QMessageBox.question(self, 'Уведомление', message, QtWidgets.QMessageBox.Ok,
                                                   QtWidgets.QMessageBox.Close)
            if reply == QtWidgets.QMessageBox.Ok:
                print('продолжить')
                self.showDialog()
            else:
                print('отмена')
        print('Открыть файл: ', self.fname)

    def nachalo(self):

        self.atribut = [self.line_zakaz.text(), self.line_project.text(), self.line_project_name.text(),
                        self.line_stage.text(), 'Solaris']

        def SS(prom):
            spisok = rashet_sekcii.rashet(self.atribut, self.fname, prom).zapusk()
            print(spisok)
            self.saveFileDialog(spisok)
            QtWidgets.QMessageBox.question(self, 'Уведомление', 'Расчет окончен!', QtWidgets.QMessageBox.Ok)
            print('конец расчетов')
            self.fname = ''

        # проверка атрибутов (номер заказа, выбран ли файл)
        if self.fname == '':  # ничего не выбрано показывает предупреждение
            print('файл не выбран')
            self.showDialog()
        elif self.atribut[0] == '':
            print(self.atribut)
            # QMessageBox.question(self, 'Уведомление', 'Введите номер заказа', QMessageBox.Ok, QMessageBox.No)
            text, ok = QInputDialog.getText(self, 'Номер заказа', 'Введите номер заказа')

            if ok:
                self.atribut[0] = str(text)
                print('Введен номер заказа: ', self.atribut[0])

        if self.atribut[1] == '':
            self.atribut[1] = 'БП'

        if self.atribut[2] == '':
            self.atribut[2] = 'Без проекта'


        # проверянм оператора уверен ли он в том, что выбрал правильное проивзоство
        if self.v1button.isChecked() == True:
            reply = QMessageBox.question(self, 'Уведомление', 'Расчет для Питон Кама!', QMessageBox.Yes, QMessageBox.No)

            if reply == QMessageBox.Yes:
                print('Расчет для производственной площадки Питон Кама')
                self.atribut[4] = 'Kama'

        else:
            reply = QMessageBox.question(self, 'Уведомление', 'Расчет для Солярис!', QMessageBox.Yes, QMessageBox.No)

            if reply == QMessageBox.Yes:
                print('Рачсет для производственной площадки Солярис')
                self.atribut[4] = 'Solaris'

        # выбираем что будем запускать: создание цеховой ведомости, изменение маршрутной карты или объединение
        # нескольких цеховых ведомостей

        if self.zadacha.currentText() == 'Создать ЦВ':
            SS(self.atribut[4])
        elif self.zadacha.currentText() == 'Переработать ЦВ':

            if self.atribut[3] == '':
                print('атрибуты ', self.atribut)
                text, ok = QInputDialog.getText(self, 'Номер заказа', 'Введине номер этапа')

                if ok:
                    self.atribut[3] = str(text)
                    print('Введен номер этапа: ', self.atribut[3])

            item = tools.exl_to_list('all', 'all', self.fname, 'Маршрутная карта').reading()

            if item == False:
                QtWidgets.QMessageBox.question(self, 'Уведомление', 'Нет нужного листа в книге',
                                               QtWidgets.QMessageBox.Ok)
                print('нет нужного листа в книге')
                self.fname = ''
            else:
                item = tools.list_to_dict(item, self.zadacha.currentText()).transformation()
                Fn = tools.writeToExl(item, self.atribut, self.fname, self.zadacha.currentText()).Deleting_sheets()
                Fn = tools.writeToExl(item, self.atribut, Fn, self.zadacha.currentText()).Write()
                self.saveFileDialog2(Fn)
                QtWidgets.QMessageBox.question(self, 'Уведомление', 'Расчет окончен!', QtWidgets.QMessageBox.Ok)
                print('конец расчетов')
                self.fname = ''
        else:
            QtWidgets.QMessageBox.question(self, 'Уведомление', 'Еще не готово!', QtWidgets.QMessageBox.Ok)

    # N-значение цеховая ведомость или маршрутная карта
    def saveFileDialog2(self, Fn):
        dict_path_save = re.findall(r'[^/]+', self.fname)  # разделили путь к файлу на список
        fname = dict_path_save[-1]  # узнали имя исходного файла
        del dict_path_save[-1]  # удалили последний элемент списка
        path_save = '/'.join(dict_path_save)  # создали новый путь для сохранения
        dirlist = QFileDialog.getSaveFileName(self, "Выбрать папку", self.fname, 'XML files (*.xml *.xlsx)')  # выбор пользователем пути сохра
        print('Выбрана путь для сохранения: ', dirlist[0])
        print('Сохранить файл: ', fname)
        put = os.path.abspath(str(Fn))  # находим путь исходного файла
        print('Путь ', put)
        shutil.copyfile(put, dirlist[0])  # копируем исходный файл в нужную дирректорию
        print('удаляем фаил ', put)
        os.remove(str(put))

    def saveFileDialog(self, spisok):
        # алгоритм сохранени в нужной папке и удалением исходника
        # i-имя файла, fileName-путь до файла
        def SaveFile(i, fileName, znach):
            print('Сохранить файл: ', fileName)
            put = os.path.abspath(str(i))  # находим путь исходного файла
            print('Путь ', put)
            shutil.copyfile(put, fileName)  # копируем исходный файл в нужную дирректорию

            if znach == 0:
                print('удаляем фаил ', i)
                os.remove(str(i))
            else:
                print('не удаляем фаил ', i)

        dict_path_save = re.findall(r'[^/]+', self.fname)  # разделили путь к файлу на список
        fname = dict_path_save[-1]  # узнали имя исходного файла
        del dict_path_save[-1]  # удалили последний элемент списка
        path_save = '/'.join(dict_path_save)  # создали новй путь для сохранения
        dirlist = QFileDialog.getExistingDirectory(self, "Выбрать папку", path_save)  # выбор пользователем пути сохра
        print('Выбрана папка для сохранения: ', dirlist)
        print('SPISOK\n', spisok)
        dirlistNew = str(dirlist) + '/№' + str(self.atribut[0] +
                                               ' Проект №' + str(self.atribut[1]) +
                                               ' ' + str(self.atribut[2]))

        # проверяем есть ли создаваемая папка
        def proverka(fileP):
            if os.path.exists(fileP):
                print('папка уже существует')
            else:
                print('Записываем новую папку')
                os.mkdir(fileP)

        # создаем стандартные папки и запускаем расчет материалов.
        proverka(dirlistNew)            # проверяем наличие папок
        # начинаем создавать свои папки для файлов
        spisokPapok = ['/1. Ведомость деталей/', '/2. КОВ/', '/3. Наклейки/',
                       '/4. Программы/', '/5. Расходники/', '/6. Чертежи/', '/7. Материалы/',
                       '/8. Спецификация заказа/']

        # проверяем наличие папок
        for i in spisokPapok:
            proverka(dirlistNew + i)

        # перезаписываем файлы созданные при расчете ЦВ
        for i in spisok[0]:
            print('Файл:', i)
            fileName = dirlistNew + spisokPapok[i[0]] + str(i[1])  # создаем новый путь для сохранения файла
            SaveFile(i[1], fileName, 0)

        for i in spisok[1]:
            rezult = materials_simple.vvod(spisok[1][i]['seria'], spisok[1][i]['material'],
                                           spisok[1][i]['nominal'], spisok[1][i]['dlina'],
                                           spisok[1][i]['Nstik'], spisok[1][i]['Nsekc'], spisok[1][i]['Nkon_zag'],
                                           spisok[1][i]['Nflanc'], spisok[1][i]['Lsvar_izd'],
                                           4).fg()
            print('Преобразование', rezult)
            to_exl_main_mat.preobrazovanie(rezult, [spisok[1][i]['seria'], spisok[1][i]['material'],
                                                    spisok[1][i]['nominal'], spisok[1][i]['dlina'],
                                                    spisok[1][i]['Nstik'], spisok[1][i]['Nsekc'],
                                                    spisok[1][i]['Nkon_zag'],
                                                    spisok[1][i]['Nflanc'], spisok[1][i]['Lsvar_izd']])
            namefile = 'Расчет маетриалов Заказ №' + str(self.line_zakaz.text()) + ' Этап №' + str(
                i) + ' ' + str(spisok[1][i]['seria']) + '-' + str(spisok[1][i]['material']) + '-' + str(
                spisok[1][i]['nominal']) + '-' + str(spisok[1][i]['dlina'])
            SaveFile("example2.xls", dirlistNew + spisokPapok[6] + namefile + '.xls', 0)
            print('fname', fname, '\n', self.fname)
            SaveFile(str(self.fname), dirlistNew + spisokPapok[7] + fname, 1)

# расчет материалов
class ExampleApp2(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__(parent)
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(340, 600)
        self.layoutWidget = QtWidgets.QWidget(Dialog)
        self.layoutWidget.setGeometry(QtCore.QRect(50, 20, 249, 520))
        self.layoutWidget.setObjectName("layoutWidget")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.layoutWidget)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.label_1 = QtWidgets.QLabel(self.layoutWidget)
        self.label_1.setObjectName("label")
        self.verticalLayout.addWidget(self.label_1)
        self.seria = QtWidgets.QComboBox(self.layoutWidget)
        self.seria.setObjectName("seria")
        self.seria.addItem("")
        self.seria.addItem("")
        self.verticalLayout.addWidget(self.seria)

        self.label_0 = QtWidgets.QLabel(self.layoutWidget)
        self.label_0.setObjectName("label")
        self.verticalLayout.addWidget(self.label_0)

        self.material = QtWidgets.QComboBox(self.layoutWidget)
        self.material.setObjectName("material")
        self.material.addItem("")
        self.material.addItem("")
        self.verticalLayout.addWidget(self.material)
        self.label = QtWidgets.QLabel(self.layoutWidget)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.nominal = QtWidgets.QComboBox(self.layoutWidget)
        self.nominal.setObjectName("nominal")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.nominal.addItem("")
        self.verticalLayout.addWidget(self.nominal)

        self.label_8 = QtWidgets.QLabel(self.layoutWidget)
        self.label_8.setObjectName("label_8")
        self.verticalLayout.addWidget(self.label_8)
        self.KolProv = QtWidgets.QComboBox(self.layoutWidget)
        self.KolProv.setObjectName("kolprov")
        self.KolProv.addItem("")
        self.KolProv.addItem("")
        self.KolProv.addItem("")
        self.verticalLayout.addWidget(self.KolProv)

        self.label_2 = QtWidgets.QLabel(self.layoutWidget)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2)
        self.dlina_trassi = QtWidgets.QLineEdit(self.layoutWidget)
        self.dlina_trassi.setObjectName("dlina_trassi")
        self.verticalLayout.addWidget(self.dlina_trassi)
        self.label_3 = QtWidgets.QLabel(self.layoutWidget)
        self.label_3.setObjectName("label_3")
        self.verticalLayout.addWidget(self.label_3)
        self.kol_stik = QtWidgets.QLineEdit(self.layoutWidget)
        self.kol_stik.setObjectName("kol_stik")
        self.verticalLayout.addWidget(self.kol_stik)
        self.label_4 = QtWidgets.QLabel(self.layoutWidget)
        self.label_4.setObjectName("label_4")
        self.verticalLayout.addWidget(self.label_4)
        self.kol_seks = QtWidgets.QLineEdit(self.layoutWidget)
        self.kol_seks.setObjectName("kol_seks")
        self.verticalLayout.addWidget(self.kol_seks)
        self.label_5 = QtWidgets.QLabel(self.layoutWidget)
        self.label_5.setObjectName("label_5")
        self.verticalLayout.addWidget(self.label_5)
        self.kol_zag = QtWidgets.QLineEdit(self.layoutWidget)
        self.kol_zag.setObjectName("kol_zag")
        self.verticalLayout.addWidget(self.kol_zag)
        self.label_6 = QtWidgets.QLabel(self.layoutWidget)
        self.label_6.setObjectName("label_6")
        self.verticalLayout.addWidget(self.label_6)
        self.kol_flanc = QtWidgets.QLineEdit(self.layoutWidget)
        self.kol_flanc.setObjectName("kol_flanc")
        self.verticalLayout.addWidget(self.kol_flanc)
        self.label_7 = QtWidgets.QLabel(self.layoutWidget)
        self.label_7.setObjectName("label_7")
        self.verticalLayout.addWidget(self.label_7)
        self.dlina_svarka = QtWidgets.QLineEdit(self.layoutWidget)
        self.dlina_svarka.setObjectName("dlina_svarka")
        self.verticalLayout.addWidget(self.dlina_svarka)
        self.verticalLayout_2.addLayout(self.verticalLayout)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.pushButton = QtWidgets.QPushButton(self.layoutWidget)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        self.pushButton2 = QtWidgets.QPushButton(self.layoutWidget)
        self.pushButton2.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton2)
        self.verticalLayout_2.addLayout(self.horizontalLayout)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

        btn = self.pushButton  # если кнопка нажата запускаем расчет материадлов
        btn.clicked.connect(self.zapusk)

        btn2 = self.pushButton2  # если кнопка нажата запускаем объединение файлов
        btn2.clicked.connect(self.slogenie)
        self.NameDIr = 'D:/Загрузки'
        self.fileName = ''

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Расчет материалов"))
        self.seria.setItemText(0, _translate("Dialog", "E3"))
        self.seria.setItemText(1, _translate("Dialog", "CR1"))
        self.material.setItemText(0, _translate("Dialog", "Алюминий"))
        self.material.setItemText(1, _translate("Dialog", "Медь"))
        self.label.setText(_translate("Dialog", "Номинал тока А."))
        self.nominal.setItemText(0, _translate("Dialog", "630"))
        self.nominal.setItemText(1, _translate("Dialog", "800"))
        self.nominal.setItemText(2, _translate("Dialog", "1000"))
        self.nominal.setItemText(3, _translate("Dialog", "1250"))
        self.nominal.setItemText(4, _translate("Dialog", "1600"))
        self.nominal.setItemText(5, _translate("Dialog", "2000"))
        self.nominal.setItemText(6, _translate("Dialog", "2500"))
        self.nominal.setItemText(7, _translate("Dialog", "3200"))
        self.nominal.setItemText(8, _translate("Dialog", "4000"))
        self.nominal.setItemText(9, _translate("Dialog", "5000"))
        self.nominal.setItemText(10, _translate("Dialog", "6400"))
        self.KolProv.setItemText(0, _translate("Dialog", "3"))
        self.KolProv.setItemText(1, _translate("Dialog", "4"))
        self.KolProv.setItemText(2, _translate("Dialog", "5"))
        self.label_0.setText(_translate("Dialog", "Материал проводника"))
        self.label_1.setText(_translate("Dialog", "Серия шинопровода"))
        self.label_2.setText(_translate("Dialog", "Длина трассы м.п."))
        self.label_3.setText(_translate("Dialog", "Количество стыков шт."))
        self.label_4.setText(_translate("Dialog", "Количество секций шт."))
        self.label_5.setText(_translate("Dialog", "Количество концевых заглушек шт."))
        self.label_6.setText(_translate("Dialog", "Количество фланцевых блоков шт."))
        self.label_7.setText(_translate("Dialog", "Длина сварных изделий м.п."))
        self.label_8.setText(_translate("Dialog", "Количество проводников"))
        self.pushButton.setText(_translate("Dialog", "Запустить"))
        self.pushButton2.setText(_translate("Dialog", "Объеденить файлы"))

    def saveFileDialog(self):
        if self.fileName == '':
            self.fileName = 'Файл сложения'
        else:
            print('Итоговое имя файлу задано: ', self.fileName)
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        self.NameDIr = QFileDialog.getExistingDirectory(self, "Выбрать папку", self.NameDIr)
        print(self.NameDIr)
        put = os.path.abspath('example2.xls')
        print(put)
        shutil.copyfile(put, self.NameDIr + '/' + str(self.fileName))
        os.remove('example2.xls')

    def zapusk(self):
        print('запускаем расчет материалов')
        self.fileName = 'Материалы ' + str(self.seria.currentText()
                                           + '-' + str(self.material.currentText())
                                           + '-' + str(self.nominal.currentText())
                                           + '-' + str(self.dlina_trassi.text()) + 'м.п.xls')
        rezult = materials_simple.vvod(self.seria.currentText(), self.material.currentText(),
                                       self.nominal.currentText(), self.dlina_trassi.text(),
                                       self.kol_stik.text(), self.kol_seks.text(), self.kol_zag.text(),
                                       self.kol_flanc.text(), self.dlina_svarka.text(), self.KolProv.currentText()).fg()
        print('преобразование')
        to_exl_main_mat.preobrazovanie(rezult, [self.seria.currentText(), self.material.currentText(),
                                                self.nominal.currentText(), self.dlina_trassi.text(),
                                                self.kol_stik.text(), self.kol_seks.text(), self.kol_zag.text(),
                                                self.kol_flanc.text(), self.dlina_svarka.text()])


        self.saveFileDialog()

    def slogenie(self):
        '''модуль сложения нескольких файлов в один'''
        zse = 0
        vals = []

        while zse != 1:
            fname = QFileDialog.getOpenFileName(self, 'Открыть файл', '', "Exel (*.xls)", )[0]
            print('Открыть файл: ', fname)
            rb = xlrd.open_workbook(fname, on_demand=True, formatting_info=True)
            nameSheet = rb.sheet_names()  # считываем имена листов в файле эксель
            nomSheet = nameSheet.index('Артикулы')  # ищем номер нужного листа и записываем его
            sheet = rb.sheet_by_index(nomSheet)  # открываем нужный нам лист
            vals_one = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]  # делаем список из всех строк
            rb.release_resources()
            del rb
            del vals_one[0]
            vals = vals + vals_one
            zse = self.vopros(vals)

    def vopros(self, vals):
        message = 'Открыть другой файл?'
        reply = QtWidgets.QMessageBox.question(self, 'Уведомление', message, QtWidgets.QMessageBox.No,
                                               QtWidgets.QMessageBox.Yes)
        if reply == QtWidgets.QMessageBox.No:
            print('Ответ НЕТ')
            itog = sorting_det(vals)
            for i in vals:
                print(i)
            print('Список отсортирован')
            to_exl_main_mat.preobrazovanie(itog, ['Суммирование'])
            print('сохранение файла')
            self.saveFileDialog()
            zse = 1
            return zse
        else:
            print('Продолжить расчеты')
            zse = 0
            return zse


def sorting_det(array):
    array2 = []
    kol_str = len(array)  # записываем клличесвто элементов в списке, оно же
    print('всего строк ', kol_str)

    for x in array:  # пока Х в списке
        kol = float(x[3])  # количество деталей из строки
        print('Строка которую сравниваем: ', x)
        y = array.index(x) + 1  # нам нужна следующая строка
        print(y, 'YYYYYYYYY')
        sov = 0  # это счетчик совпадений
        while y <= kol_str - 1:  # начинаем сравнивать со следующей строкой
            w = array[y]  # строка с которой сравниваем это следующая строка
            print('Строка c которой сравниваем: ', w)
            if x[0] == w[0]:  # если значения строк совпадают, то
                kol = kol + float(w[3])  # складываем количество
                sov = sov + 1  # сообщаем что были совпадения
                print('*' * 25 + 'СОВПАДЕНИЕ!!!' + '*' * 25)
                print('Совпадение в ' + str(y) + ' строке  +' + str(w[3]) + 'шт.')
                print('Строка которую удаляем: ', array[y])
                del array[y]
                kol_str = len(array)
                print('Итого строк осталось: ', kol_str)
            else:
                print('нет совпадений в ', y, ' строке')
                y = y + 1

        # проверяем есть ли совпадения

        if sov == 0:
            array2.append(x)
            print('***Совпадений нет!***\n')
        else:
            spis = [x[0], x[1], x[2], str(kol), x[4], x[5]]
            print('Записываем в список ', spis)
            array2.append(spis)
        print('КОЛИЧЕСТВО СОВПАДЕНИЙ: ' + str(sov) + '\n')

    # записываем результат в файл
    # array2.sort(key=lambda i: (i[3]))
    return array2


class ExcelProcessingApp(QMainWindow):
    def __init__(self, parent=None):
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__(parent)
        self.initUI(self)  # Это нужно для инициализации нашего дизайна


    def initUI(self, discrepancies):
        discrepancies.setWindowTitle("Excel Processing App")
        discrepancies.setGeometry(100, 100, 400, 200)
        self.layout = QVBoxLayout()

        self.btn_select_made = QPushButton("Изготовлено")
        self.btn_select_made.clicked.connect(self.select_made_file)
        self.layout.addWidget(self.btn_select_made)

        self.made_file_label = QLabel("Выберите файл *xlsx")
        self.layout.addWidget(self.made_file_label)

        self.btn_select_to_make = QPushButton("Нужно изготовить")
        self.btn_select_to_make.clicked.connect(self.select_to_make_file)
        self.layout.addWidget(self.btn_select_to_make)

        self.to_make_file_label = QLabel("Выберите файл *xlsx")
        self.layout.addWidget(self.to_make_file_label)

        self.btn_run_processing = QPushButton("Пуск")
        self.btn_run_processing.clicked.connect(self.run_processing)
        self.layout.addWidget(self.btn_run_processing)

        self.central_widget = QWidget()
        self.central_widget.setLayout(self.layout)
        self.setCentralWidget(self.central_widget)

        self.made_filename = ""
        self.to_make_filename = ""

    def select_made_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_dialog = QFileDialog.getOpenFileName(self, "Выберите файл изготовленного", "", "Excel Files (*.xlsx)", options=options)
        if file_dialog[0]:
            self.made_filename = file_dialog[0]
            self.made_file_label.setText(os.path.basename(self.made_filename))

    def select_to_make_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_dialog = QFileDialog.getOpenFileName(self, "Выберите файл для изготовления", "", "Excel Files (*.xlsx)", options=options)
        if file_dialog[0]:
            self.to_make_filename = file_dialog[0]
            self.to_make_file_label.setText(os.path.basename(self.to_make_filename))

    def run_processing(self):
        if not self.made_filename or not self.to_make_filename:
            QMessageBox.warning(self, "Ошибка", "Выберите оба файла")
            return

        self.process_file(self.made_filename, self.to_make_filename)

    def process_file(self, filename1, filename2):
        print(filename1, filename2)
        sheet = "Спецификация"

        setup(filename1, filename2, sheet)

        # После завершения обработки, сохраняем файл в выбранную папку и удаляем исходный файл
        output_folder = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения файла")
        if output_folder:
            today_date = QDateTime.currentDateTime().toString("ddMMyy")
            new_filename = os.path.join(output_folder, f"{today_date}.xlsx")
            shutil.move("result.xlsx", new_filename)
            QMessageBox.information(self, "Успешно", "Файл успешно сохранен и исходный файл удален")

def setup(file_path1, file_path2, sheet_name):
    data1 = read(file_path1, sheet_name)
    data2 = read(file_path2, sheet_name)
    print('СПИСОК 1\n', data1)
    print('СПИСОК 2\n', data2)
    data1 = addition_of_repetitions(data1)
    data2 = addition_of_repetitions(data2)
    # result = compare_dicts_with_counts(data1, data2)
    result1 = vendor_code(data1)
    result2 = vendor_code(data2)
    result = differences(result1, result2)
    save_exel(result[0], result[1])

def read(file_path, sheet_name):
    print('ИЗ ЭКСЕЛЬ В СЛОВАРЬ')
    # Открываем файл Excel
    file_path = file_path  # Укажите путь к вашему файлу Excel
    wb = openpyxl.load_workbook(file_path, data_only=True)  # Используем data_only=True
    sheet = wb[sheet_name]  # Выбираем страницу "Спецификация"

    # Инициализируем словарь для хранения данных
    data_dict = defaultdict(list)

    # Получаем имена столбцов (первая строка в файле Excel)
    columns = [cell.value for cell in sheet[1]]

    # Итерируемся по строкам, начиная со второй (первая строка уже использована для имен столбцов)
    for row in sheet.iter_rows(min_row=2, values_only=True, max_col=14):  # Чтение только до 14 столбца включительно
        # Пропускаем строки, где первая ячейка содержит значения "Итого", "итого", "Итого:", "итого:", "Итог", "итог", "Итог:", "итог:"
        if row[0] in ["Итого", "итого", "Итого:", "итого:", "Итог", "итог", "Итог:", "итог:"]:
            break

        # Пропускаем строки, где первая ячейка пуста
        if row[0] is None:
            continue

        # Создаем словарь данных для строки, используя имена столбцов
        row_data = {columns[i]: row[i] for i in range(min(len(columns), len(row)))}

        # Добавляем данные в словарь по соответствующим ключам
        for key, value in row_data.items():
            if key != '№ п/п':  # Пропускаем первый столбец
                data_dict[key].append(value)

    # Закрываем файл Excel
    wb.close()

    # Выводим полученный словарь
    for key, values in data_dict.items():
        print(key, values)

    return data_dict

def addition_of_repetitions(data_dict):
    print('\nСЛОЖЕНИЕ ПОВТОРОВ')
    # Создаем новый словарь для результата
    result_dict = {
        'Серия': [],
        'IP': [],
        'Мат. Пров.': [],
        'Кол. Пров.': [],
        'Ном. ток, А': [],
        'Наименование': [],
        'Обозначение': [],
        'Тип': [],
        'Размер, мм': [],
        'Номер элемента': [],
        'Кол, шт': [],
        'Примечание': [],
        'Этап': []
    }

    # Создаем словарь, который будет хранить сумму для каждой группы (группируем по ключам)
    grouped_data = {}

    # Проходимся по каждой строке в исходном словаре
    for i in range(len(data_dict['Серия'])):
        row_data = {k: data_dict[k][i] for k in data_dict}
        key = tuple(row_data[k] for k in row_data if k != 'Кол, шт')  # Создаем ключ для группировки без 'Кол, шт'

        # Если ключ уже есть в grouped_data, обновляем 'Кол, шт'
        if key in grouped_data:
            grouped_data[key]['Кол, шт'] += row_data['Кол, шт']
        else:
            grouped_data[key] = row_data

    # Преобразуем grouped_data обратно в result_dict
    for k, v in grouped_data.items():
        for key in result_dict.keys():
            result_dict[key].append(v.get(key, None))

    # Выводим полученный словарь
    print(result_dict)
    return result_dict

def compare_dicts_with_counts(dict1, dict2):
    extra_dict = {}
    missing_dict = {}

    # Создаем множества из ключей для более быстрого поиска
    keys1 = set(dict1.keys())
    keys2 = set(dict2.keys())

    # Находим лишние ключи в первом словаре
    extra_keys = keys1 - keys2
    for key in extra_keys:
        extra_dict[key] = dict1[key]

    # Находим ключи, которых не хватает в первом словаре
    missing_keys = keys2 - keys1
    for key in missing_keys:
        missing_dict[key] = dict2[key]

    # Находим различия в значениях одинаковых ключей
    common_keys = keys1.intersection(keys2)
    for key in common_keys:
        values1 = dict1[key]
        values2 = dict2[key]

        # Если значения не совпадают, добавляем их в оба словаря
        if values1 != values2:
            extra_dict[key] = values1
            missing_dict[key] = values2

        # Проверяем количество элементов
        if len(values1) != len(values2):
            extra_dict[key + "_count"] = len(values1)
            missing_dict[key + "_count"] = len(values2)

    print("Лишнее:")
    print(extra_dict)
    print("\nНе хватает:")
    print(missing_dict)
    return extra_dict, missing_dict

def vendor_code(data):
    print('\nСоздаем список уникальных ключей на основе данных')
    unique_keys = [
        f'{data["Серия"][i]}-{data["IP"][i]}-{data["Мат. Пров."][i]}-{data["Кол. Пров."][i]}-{data["Ном. ток, А"][i]}-{data["Обозначение"][i]}-{data["Тип"][i]}-{data["Размер, мм"][i]}'
        for i in range(len(data["Серия"]))]

    # Создаем словарь для подсчета уникальных ключей
    key_count = {}
    for key in unique_keys:
        if key in key_count:
            key_count[key] += 1
        else:
            key_count[key] = 1

    # Выводим результат
    # for key, value in key_count.items():
    #     print(f'{key}: {value}')
    print(key_count)

    return (key_count)

def differences(dict1, dict2):
    # Создаем словарь "лишнее"
    extra_dict = {}
    for key in dict1:
        if key in dict2:
            if dict1[key] > dict2[key]:
                extra_dict[key] = dict1[key] - dict2[key]
        else:
            extra_dict[key] = dict1[key]

    # Создаем словарь "не хватает"
    missing_dict = {}
    for key in dict2:
        if key in dict1:
            if dict2[key] > dict1[key]:
                missing_dict[key] = dict2[key] - dict1[key]
        else:
            missing_dict[key] = dict2[key]

    # Выводим результат
    print("лишнее:")
    print(extra_dict)

    print("не хватает:")
    print(missing_dict)

    return (extra_dict, missing_dict)

def save_exel(dict1, dict2):
    print('\nСоздаем новую книгу Excel')
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()

    # Создаем страницу "лишнее"
    extra_sheet = wb.create_sheet(title="лишнее")

    # Записываем заголовок для "лишнее"
    extra_sheet.cell(row=1, column=1, value="Артикул")
    extra_sheet.cell(row=1, column=2, value="Количество")

    # Записываем данные из первого словаря на "лишнее"
    row_num = 2
    for key, value in dict1.items():
        extra_sheet.cell(row=row_num, column=1, value=key)
        extra_sheet.cell(row=row_num, column=2, value=value)
        row_num += 1

    # Создаем страницу "не хватает"
    missing_sheet = wb.create_sheet(title="не хватает")

    # Записываем заголовок для "не хватает"
    missing_sheet.cell(row=1, column=1, value="Артикул")
    missing_sheet.cell(row=1, column=2, value="Количество")

    # Записываем данные из второго словаря на "не хватает"
    row_num = 2
    for key, value in dict2.items():
        missing_sheet.cell(row=row_num, column=1, value=key)
        missing_sheet.cell(row=row_num, column=2, value=value)
        row_num += 1

    # Сохраняем книгу в файл
    wb.save("result.xlsx")

def main():
    sys.stdout = logg_solaris.print_to_txt("logfilename.txt")
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    sys.exit(app.exec_())  # и запускаем приложение


if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()

# pyinstaller --onefile --icon=sol.ico --noconsole Solaris_vedomost.py
# В случае если не работает pyinstaller
# C:\Users\anton\AppData\Local\Programs\Python\Python39\Scripts\pyinstaller --icon=sol.ico Solaris_vedomost.py
# C:\Users\anton\AppData\Local\Programs\Python\Python39\Scripts\pyinstaller --onefile --icon=sol.ico Solaris_vedomost.py
# /usr/local/bin/pyinstaller --onefile --icon=sol.ico Solaris_vedomost.py
