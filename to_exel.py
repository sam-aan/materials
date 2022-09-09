#этот модуль записывает все данные их файлов .txt в файлы .exlx
# coding: utf8
import xlwt
import re
import openpyxl
from openpyxl.styles import Font
import datetime
import сравнение

class writing_to_exl():
    def __init__(self, atrib, etap, array, last_line, ved):
        self.font0 = xlwt.Font()
        self.style0 = xlwt.XFStyle()
        self.font1 = xlwt.Font()
        self.style1 = xlwt.XFStyle()
        self.font2 = xlwt.Font()
        self.style2 = xlwt.XFStyle()
        self.font3 = xlwt.Font()
        self.style3 = xlwt.XFStyle()
        self.oboznach = ["Прямая секция",
                         "Прямая секция с фланцем",
                         "Прямая секция с закрытым фланцем",
                         "Прямая распределительная секция с фикс. Выводом",
                         "Прямая распределительная секция с розеткой",
                         "Угловая горизонтальная секция",
                         "Угловая горизонтальная секция c фланцем",
                         "Угловая вертикальная секция",
                         "Угловая вертикальная секция с фланцем",
                         "Правая комбинированная секция",
                         "Правая комбинированная секция с фланцем уг",
                         "Правая комбинированная секция с фланцем ув",
                         "Левая комбинированная секция",
                         "Левая комбинированная секция с фланцем уг",
                         "Левая комбинированная секция с фланцем ув",
                         "z-образная вертикальная секция с фланцем",
                         "z-образная вертикальная секция",
                         "z-образная горизонтальная с фланцем",
                         "z-образная горизонтальная",
                         "Тройник вертикальный",
                         "Тройник вертикальный с фланцем",
                         "Тройник горизонтальный",
                         "Тройник горизонтальный с флацем",
                         "Секция компенсации",
                         "Секция перевода фаз",
                         "Секция перевода нейтрали",
                         "Блок отбора мощности",
                         "Блок отбора мощности фикированный",
                         "Трансформаторная секция вертикальная",
                         "Трансформаторная секция вертикальная с углом вертикальным",
                         "Трансформаторная секция вертикальная с углом горизонтальным",
                         "Трансформаторная секция вертикальная с тройником",
                         "Трансформаторная секция горизонтальная",
                         "Трансформаторная секция горизонтальная с углом вертикальным",
                         "Трансформаторная секция горизонтальная с углом горизонтальным",
                         "Трансформаторная секция горизонтальная с тройником",
                         "Концевая заглушка",
                         "Стыковочный блок",
                         "Огнепожарная проходка",
                         "Крепежная скоба",
                         "Пружинный подвес",
                         "Жесткий подвес",
                         "Адаптер",
                         "Крышка стыка"]
        self.oboznach_1 = ['Крышка СБ', 'Крышка с выступом СБ', 'Стенка СБ', 'Шина СБ', 'Направляющая', 'Сухарь',
                           'Фланец', 'Шина СБ (Зеркальная)']
        self.wb = openpyxl.Workbook()   # Создаем виртуальную книгу
        self.zakaz = atrib      # номер заказа
        self.etap = etap        # номер этапа
        self.last_line = last_line  # номер расчет, данные о компьютере.
        self.slovar = array
        self.ved = ved

    def setting_exl(self):
        """Настройки"""
        print('13. ЗАПИСЬ В EXLX')
        self.font0 = xlwt.Font()
        self.font0.name = 'Century Gothic'
        self.font0.colour_index = 0
        self.font0.bold = True
        self.font0.italic = True
        self.font0.underline = True

        self.style0 = xlwt.XFStyle()
        self.style0.font = self.font0

        self.font1 = xlwt.Font()
        self.font1.name = 'Century Gothic'
        self.font1.colour_index = 0

        self.style1 = xlwt.XFStyle()
        self.style1.font = self.font1

        self.font2 = xlwt.Font()
        self.font2.name = 'Century Gothic'
        self.font2.colour_index = 0
        self.font2.bold = True

        self.style2 = xlwt.XFStyle()
        self.style2.font = self.font2

        self.font3 = xlwt.Font()
        self.font3.name = 'Century Gothic'
        self.font3.colour_index = 0
        self.font3.colour_index.bit_length()
        self.font3.bold = True

        self.style3 = xlwt.XFStyle()
        self.style3.font = self.font3


        if self.ved == 'upakovka':
            print('СОЗДАНИЕ КОНТРОЛЬНОЙ ВЕДОМОСТИ')
            spis_dla_vozvrata = self.upakovka()
            #self.OTK()
            self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
            name_out_file = 'Контрольная ведомость Заказ №' + str(self.zakaz) + '.xlsx'
            print('Сохранение в файл: ', name_out_file)
            self.wb.save(name_out_file)
            return [[1, name_out_file], spis_dla_vozvrata]
        else:
            self.Nach_smen()
            self.txt_tabl()
            #self.komlekt_ie()
            self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))
            name_out_file = 'Заказ №' + str(self.zakaz) + ' Этап №' + str(self.etap) + '.xlsx'
            print('Сохранение в файл: ', name_out_file)
            self.wb.save(name_out_file)
            return [0, name_out_file]

    def kolontitul(self, name_sheet):
        """колонтитулы"""
        # верхний левый колонтитул
        sheet = self.wb[name_sheet]
        if self.ved == 'upakovka':
            sheet.oddHeader.left.text = 'Заказ № ' + str(self.zakaz)
        else:
            sheet.oddHeader.left.text = 'Заказ № ' + str(self.zakaz) + '\nЭтап № ' + str(self.etap)
        sheet.oddHeader.left.size = 12
        sheet.oddHeader.left.font = "Century Gothic"

        # верхний центральный колонтитул
        sheet.oddHeader.center.text = name_sheet
        sheet.oddHeader.center.size = 20
        sheet.oddHeader.center.font = "Century Gothic"

        # верхний правый колонтитул
        now = datetime.datetime.today().strftime("%d %m %Y")
        sheet.oddHeader.right.text = 'ООО «ПИК «СОЛЯРИС»\n' + str(now)
        sheet.oddHeader.right.size = 12
        sheet.oddHeader.right.font = "Century Gothic"

        # нижний правый колонтитул
        now = datetime.datetime.today().strftime("%d %m %Y")
        sheet.oddFooter.right.text = str(self.last_line)
        sheet.oddFooter.right.size = 10
        sheet.oddFooter.right.font = "Century Gothic"

        if name_sheet in ['пила', 'гибка', 'фрезер', 'сверловка', 'напыление', 'сварка', 'зачистка',
                          'шлифовка', 'покраска', 'предсборка', 'сборка']:
            # нижний левый колонтитул
            sheet = self.wb[name_sheet]
            sheet.oddFooter.left.text = 'Время начала:____:____' + '\n' + 'Время завершения:____:____'
            sheet.oddFooter.left.size = 14
            sheet.oddFooter.left.font = "Century Gothic"

            # нижний центральный колонтитул
            sheet.oddFooter.center.text = 'ФИО исполнителя:____________'
            sheet.oddFooter.center.size = 14
            sheet.oddFooter.center.font = "Century Gothic"

    def isFlot(self, value):
        """проверяем является ли значение числом"""
        try:
            float(value)
            return True
        except ValueError:
            return False

    def Nach_smen(self):

        """модуль записывает в файл эксель данные для нач. смены и раб. мест."""

        print('создана страница для начальника смены')
        self.wb.create_sheet(title='Ведомость Нач.смен', index=1)  # создаем первый лист
        sheet = self.wb['Ведомость Нач.смен']
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        self.kolontitul('Ведомость Нач.смен')
        x = 1
        operation = ['05', '10', '15', '20', '25', '30', '35', '40', '45', '50', '55', '60']
        # записываем первую строку для ведомости нач. смены с раб. местами.
        perv_stroka = ['№', 'Серия', 'ip', 'Материал', 'In', 'Кол. пров.', 'Наименование', 'Обозначение',
                            'Разм.L', 'Разм.L1', 'Разм.A', 'Разм.B', 'Разм.C', 'Кол.',
                            'пила', 'гибка', 'фрезер', 'сверловка', 'напыление', 'сварка',
                            'зачистка', 'шлифовка', 'лаз.резка', 'покраска', 'предсборка', 'сборка_заливка']
        for i in perv_stroka:
            cell = sheet.cell(row=1, column=x)
            cell.font = Font(name='Century Gothic', bold=True, color='000000')
            cell.value = i
            x += 1

        print('созданы страницы для рабочих мест')
        rab_mesto = {'пила': 'O', 'гибка': 'P', 'фрезер': 'Q', 'сверловка': 'R', 'напыление': 'S', 'сварка': 'T',
                     'зачистка': 'U', 'шлифовка': 'V', 'лаз.резка': 'W', 'покраска': 'X', 'предсборка': 'Y',
                     'сборка_заливка': 'Z'}

        # Создаем новые страницы для раб. мест
        for s in rab_mesto:
            nomer_s = perv_stroka.index(s)  # находим порядковый номер для s
            self.wb.create_sheet(title=s, index=nomer_s)  # создаем лист
            sheet2 = self.wb[s]  # создаем новую страницу
            sheet2.sheet_properties.pageSetUpPr.fitToPage = True
            sheet2.page_setup.fitToHeight = False
            self.kolontitul(s)
            x = 1

            # записываем первую строку в новую страницу без рабочих мест
            for i in perv_stroka:

                if x <= 14:
                    cell = sheet2.cell(row=1, column=x)
                    cell.font = Font(name='Century Gothic', bold=True, color='000000')
                    cell.value = str(i)
                    x += 1
                else:
                    continue

        # начинаем заполнять ячейки для рабочих мест в ведомости начальников смен
        y = 2

        for i in self.slovar:
            x = 2

            # для обозначения секций жирным шрифтом с подчеркиванием для сборок
            if self.slovar[i]['Наименование'] in self.oboznach:
                self.slovar[i]['Обозначение'] = str(self.slovar[i]['Обозначение']) + str(self.slovar[i]['тип'])
                st = Font(name='Century Gothic', bold=True, color='000000')
            # для обозначения жирным шрифтом для обозначения подсборок
            elif self.slovar[i]['Наименование'] in self.oboznach_1:
                st = Font(name='Century Gothic', bold=True, italic=True, color='000000')
            # для обозначения деталей простым шрифтом
            else:
                st = Font(name='Century Gothic')

            # записываем номер строки в ведомость нач. смены.
            cell = sheet.cell(row=y, column=1)
            cell.font = st
            cell.value = i

            # начинаем вычислять какая строчка нам нужна для формулы значения
            nS_schet = 1                                        # номер счетчика

            for cell in self.wb['Ведомость Нач.смен']['A']:     # перебираем столбец А в списке нач смен
                #print(cell.value)
                #print('i', i)
                if cell.value == i:                             # ищем нужную сточку
                    znchStr = nS_schet                          # присваиваем значение нужной строчки переменной
                nS_schet += 1

            for j in self.slovar[i]:
                if j in ['номер элемента', 'тип', 'стандарт', 'категория']:
                    continue    # пропускаем ненужные колонки
                cell = sheet.cell(row=y, column=x)
                cell.font = st

                if self.slovar[i][j] == '0' and j in rab_mesto:
                    cell.value = 'Х'
                elif self.slovar[i][j] in operation and j in rab_mesto:     # проверяем, есть ли номер операции в списке
                    cell.value = ''
                    stroka = [i, self.slovar[i]['Серия'], self.slovar[i]['ip'], self.slovar[i]['Материал'],
                              self.slovar[i]['In'], self.slovar[i]['Кол. пров.'],
                              self.slovar[i]['Наименование'],
                              self.slovar[i]['Обозначение'], self.slovar[i]['Разм.L'], self.slovar[i]['Разм.L1'],
                              self.slovar[i]['Разм.A'], self.slovar[i]['Разм.B'], self.slovar[i]['Разм.C'],
                              '''='Ведомость Нач.смен'!N''' + str(znchStr) + '''-'Ведомость Нач.смен'!'''
                              + rab_mesto[j] + str(znchStr)]

                    xxx = 1
                    yyy = self.wb[j].max_row + 1

                    for s in stroka:
                        cell2 = self.wb[j].cell(row=yyy, column=xxx)
                        cell2.font = Font(name='Century Gothic')
                        cell2.value = str(s)
                        xxx += 1
                else:
                    cell.value = self.slovar[i][j]
                x += 1
            y = y + 1

    def upakovka(self):
        print('УПАКОВКА ОТК')
        spis_dla_vozvrata = []
        name_sheet = 'Упаковка'
        self.wb.create_sheet(title=name_sheet, index=18)  # создаем первый лист
        w_rab_mesto = self.wb[name_sheet]  # создаем новую страницу
        w_rab_mesto.sheet_properties.pageSetUpPr.fitToPage = True
        w_rab_mesto.page_setup.fitToHeight = False
        self.kolontitul(name_sheet)

        # ves E3
        vesEAl = {630: 7.6, 800: 8.8, 1000: 10.8, 1250: 13.2, 1600: 17.1, 2000: 20.3, 2500: 24.5, 3200: 32, 4000: 38.3,
                5000: 46.6, 6300: 110.25}
        vesECu = {800: 15.9, 1000: 17.8, 1250: 24.8, 1600: 31.8, 2000: 43.9, 2500: 57.1, 3200: 63.1, 4000: 87.2}
        vesStikAl = {630: 2, 800: 2.2, 1000: 2.6, 1250: 3.1, 1600: 4.4, 2000: 5.1, 2500: 5.17, 3200: 8.35}
        vesStikCu = {1000: 3.61, 1600: 5.53, 2500: 9.34, 3200: 9.36, 4000: 14.69}

        # записываем первую строку
        x = 1
        for i in ['№ п/п', 'Артикул', 'Серийный номер', 'Наименование', 'Размер, мм', 'Номер элемента', 'Отметка ОТК',
                  'Место', 'Масса, кг.', 'Примечание']:
            cell = w_rab_mesto.cell(row=1, column=x)
            cell.font = Font(name='Century Gothic', bold=True, color='000000')
            cell.value = i
            x += 1

        y = 2   # счетчик строк
        x1 = 1  # номер строки
        serNom = 1

        for i in self.slovar:
            spis_dla_vozvrata_2 = []
            d = re.findall(r'\d+', self.slovar[i]['Размер, мм'])

            if self.slovar[i]['Обозначение'] in ['УВ', 'УГ', 'УВФ', 'УГФ', 'uv', 'ug', 'uvf', 'ugf']:
                dlina = int(d[0]) + int(d[1])
            elif self.slovar[i]['Обозначение'] in ['ЗВ', 'ЗГ', 'КП', 'КЛ', 'ЗВФ', 'ЗГФ', 'ТВ', 'ТГ',
                                                   'zb', 'zg', 'kp', 'kl', 'zvf', 'zgf', 'tv', 'tg']:
                dlina = int(d[0]) + int(d[1]) + int(d[2])
            elif self.slovar[i]['Обозначение'] in ['П', 'ПФ', 'ОМ', 'ОМФ', 'ТС', 'pt', 'pf', 'pr', 'prf', 'ts']:
                dlina = int(d[0])
            else:
                dlina = 0

            # Вес шинопровода
            sss = ['УВ', 'УГ', 'УВФ', 'УГФ', 'ЗВ', 'ЗГ', 'КП', 'КЛ', 'ЗВФ', 'ЗГФ', 'ТВ', 'ТГ', 'П', 'ПФ',
                   'ОМ', 'ОМФ', 'ТС', 'uv', 'ug', 'uvf', 'ugf', 'zv', 'zg', 'kp', 'kl', 'zvf', 'zgf', 'tv', 'tg',
                   'pt', 'pf', 'pr', 'prf']
            if self.slovar[i]['Серия'] == 'E3' and self.slovar[i]['Обозначение'] in sss:

                if self.slovar[i]['Мат. Пров.'] == 'Al' and int(self.slovar[i]['Ном. ток, А']) in vesEAl:
                    ves = round(dlina * vesEAl[int(self.slovar[i]['Ном. ток, А'])] / 1000, 2)
                elif self.slovar[i]['Мат. Пров.'] == 'Cu' and int(self.slovar[i]['Ном. ток, А']) in vesECu:
                    ves = round(dlina * vesECu[int(self.slovar[i]['Ном. ток, А'])] / 1000, 2)
                else:
                    ves = '-'

            # Вес стыков
            elif self.slovar[i]['Серия'] == 'E3' and self.slovar[i]['Обозначение'] in ['СБ', 'sb']:
                if self.slovar[i]['Мат. Пров.'] == 'Al' and int(self.slovar[i]['Ном. ток, А']) in vesStikAl:
                    ves = round(vesStikAl[int(self.slovar[i]['Ном. ток, А'])], 2)
                elif self.slovar[i]['Мат. Пров.'] == 'Cu' and int(self.slovar[i]['Ном. ток, А']) in vesStikCu:
                    ves = round(vesStikCu[int(self.slovar[i]['Ном. ток, А'])], 2)
                else:
                    ves = '-'
            else:
                ves = '-'

            #   1 stolbec № п/п
            cell = w_rab_mesto.cell(row=y, column=1)
            cell.font = Font(name='Century Gothic')
            cell.value = x1

            # 2 столбец Артикул
            cell = w_rab_mesto.cell(row=y, column=2)
            cell.font = Font(name='Century Gothic')
            value = str(self.slovar[i]['Серия']) + '-' + str(self.slovar[i]['IP']) + '-' + str(
                self.slovar[i]['Мат. Пров.']) + '-' + str(self.slovar[i]['Кол. Пров.']) + '-' + str(
                self.slovar[i]['Ном. ток, А']) + '-' + str(self.slovar[i]['Обозначение'] + str(self.slovar[i]['Тип']))
            cell.value = value
            spis_dla_vozvrata_2.append(value)      # для возвращяемого списка

            #   3 stolbec Серийный номер
            if self.slovar[i]['Наименование'] in ['Крепежная скоба', 'Крышка стыка', 'Пружинный подвес']:
                value = ''
            else:
                value = str(re.findall(r'\d+', str(self.zakaz).lower())[0]) \
                        + '-' + str(re.findall(r'\d+', str(self.slovar[i]['Этап']).lower())[0]) \
                        + '-' + str("%03d" % serNom)
            cell = w_rab_mesto.cell(row=y, column=3)
            cell.font = Font(name='Century Gothic')
            cell.value = value
            serNom += 1
            spis_dla_vozvrata_2.append(value)      # для возвращяемого списка

            # 4 столбец Наименование
            cell = w_rab_mesto.cell(row=y, column=4)
            cell.font = Font(name='Century Gothic')
            cell.value = self.slovar[i]['Наименование']
            spis_dla_vozvrata_2.append(self.slovar[i]['Наименование'])      # для возвращяемого списка

            # 5 столбец Размер
            cell = w_rab_mesto.cell(row=y, column=5)
            cell.font = Font(name='Century Gothic')
            cell.value = self.slovar[i]['Размер, мм']
            spis_dla_vozvrata_2.append(self.slovar[i]['Размер, мм'])      # для возвращяемого списка

            # 6 столбец номер элемента
            cell = w_rab_mesto.cell(row=y, column=6)
            cell.font = Font(name='Century Gothic')
            cell.value = self.slovar[i]['Номер элемента']
            spis_dla_vozvrata_2.append(self.slovar[i]['Номер элемента'])      # для возвращяемого списка

            # 9 столбец масса
            cell = w_rab_mesto.cell(row=y, column=9)
            cell.font = Font(name='Century Gothic')

            if self.slovar[i]['Наименование'] in ['Крепежная скоба', 'Крышка стыка', 'Пружинный подвес']:
                cell.value = str(self.slovar[i]['количество']) + ' шт.'
                spis_dla_vozvrata_2.append(str(self.slovar[i]['количество']) + ' шт.')      # для возвращяемого списка
            else:
                cell.value = ves
                spis_dla_vozvrata_2.append(ves)      # для возвращяемого списка

            # 10 столбец примечание
            cell = w_rab_mesto.cell(row=y, column=10)
            cell.font = Font(name='Century Gothic')
            if self.slovar[i]['Обозначение'] in ['kz', 'op', 'ks', 'pp', 'gp', 'ad', 'ksb']:
                cell.value = str(self.slovar[i]['Примечание']) + str(self.slovar[i]['Кол, шт']) + ' шт.'
            else:
                cell.value = self.slovar[i]['Примечание']

            x1 += 1
            y += 1
            spis_dla_vozvrata.append(spis_dla_vozvrata_2)  # для возвращяемого списка

        return spis_dla_vozvrata

    def komlekt_ie(self):
        """записываем материалы"""
        print('Комплектующие')
        self.wb.create_sheet(title='Комплектующие', index=14)  # создаем первый лист
        sheet = self.wb['Комплектующие']
        self.kolontitul('Комплектующие')
        сравнение.sorting('txt', 'комплектующие.txt', 'комплектующие ИТОГ.txt', 4, False, ';').addition()

        tr = open('комплектующие ИТОГ.txt', 'r')  # отркываем и читаем файл
        line = (tr.readline().rstrip())  # начинаем читать построчно
        y = 1

        while line:
            q = re.findall(r'[^;]+', line)  # создаем список из строки
            x = 1

            for i in q:
                st = Font(name='Century Gothic', color='000000')
                cell = sheet.cell(row=y, column=x)
                cell.font = st
                cell.value = i
                x = x + 1
            y = y + 1
            line = tr.readline()  # присваем значению line следующую строку

    def txt_tabl(self):
        """Выводим на печать таблицу и записываем ее в текстовый файл"""
        print('СОХРАНЕНИЕ ТАБЛИЦЫ')
        #tab_spis_det = open('Таблица заказ № ' + str(self.zakaz) + 'Этап № ' + str(self.etap) + '.txt', 'w')
        s = self.slovar

        for i in s:
            a =[]
            a.append(i)
            for j in self.slovar[i]:
                a.append(self.slovar[i][j])
            tab_q = '{0:<5}{1:<5}{2:<5}{3:<5}{4:<15}{5:<5}{6:<35}{7:<15}{8:<15}' \
                '{9:<15}{10:<10}{11:<10}{12:<10}{13:<10}{14:<10}{15:<10}'.format(a[0], a[1], a[2], a[3], a[4], a[5],
                                                                                a[6], a[7], a[8], a[9], a[10], a[11],
                                                                                a[12], a[13], a[14], a[15])
            print(tab_q)
            #tab_spis_det.write(str(tab_q) + '\n')
        #tab_spis_det.close()

if __name__ == '__main__':      # Если мы запускаем файл напрямую, а не импортируем
    print('hi')
    writing_to_exl('100', '2').setting_exl()
    writing_to_exl('170', '1')