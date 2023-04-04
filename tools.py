# из эксель в словарь словарей или файл
# column    количество столбцов которое нужно захватить
# rows      количество сторк котторе нужно захватить all - если все до первой пустой строки
# inFile    путь к файлу с именем только *.xlsx
#nameStr    имя страницы которую нужно преобразовать в список списоков
# coding: utf8
import datetime
import openpyxl
import copy
from openpyxl.styles import Font


class exl_to_list:
    def __init__(self, column, rows, inFile, nameStr):
        self.col = column       # номер столбца до которой нужно производить чтение
        self.rows = rows        # номер строчки до которой нужно производить чтение
        self.inf = inFile       # путь к файлу
        self.NStr = nameStr     # Имя страницы которую нужно прочитать
        self.array = []         # пустой список, который будем возвращять

    def reading(self):
        """модуль создания из файла эксель раскрытого списка"""
        print('ЭКСЕЛЬ РАСКРЫТЫЙ СПИСКОК')
        wb = openpyxl.load_workbook(self.inf, data_only=True)

        # проверяем есть ли в книге лист с заданным названием
        if str(self.NStr) not in wb:
            return False
        else:
            sheet = wb.get_sheet_by_name(str(self.NStr))

            if self.rows in ['all']:
                rows = sheet.max_row + 1
            else:
                rows = self.rows + 1

            if self.col in ['all']:
                cols = sheet.max_column + 1
            else:
                cols = self.col + 1

            for i in range(1, rows):
                sel = []

                for j in range(1, cols):
                    cell = sheet.cell(row=i, column=j)
                    if cell.value in [None]:
                        sel.append('')
                    else:
                        sel.append(str(cell.value))

                self.array.append(sel)

            print('Возращаем список')
            return self.array

# создаем из списка, который получили из экселя словарь. каждое значение будет страницей
class list_to_dict:
    def __init__(self, list, zadacha):
        self.li = list
        self.alfavit = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                   'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF',
                   'AG', 'AH', 'AI']
        self.index = self.li[0].index('Кол.')    # находим крайний столбец
        self.zadacha = zadacha

    def transformation(self):
        """модуль создания из словаря для всех странциц маршрутной карты, что бы можно было записать в эксель"""
        print('СЛОВАРЬ ДЛЯ ВСЕХ СТРАНИЦ')

# создаем первый список
        def one_list():
            s1 = []
            for i in self.li:
                a = 0
                s11 = []
                for j in i:
                    if a <= self.index:
                        s11.append(j)
                    a += 1
                s1.append(s11)
            #print('первая половина списка\n', s1)
            return s1

# создаем словарь из всей таблицы
        def two_dict():
            s = {}
            li = copy.copy(self.li)
            for i in li[0]:
                s[str(i)] = []
            li.pop(0)
            n = 0
            for i in s.keys():
                for j in li:
                    s[i].append(j[n])
                n += 1
            #print('словарь всего списка\n', s)
            return s

# теперь мы создадим словарь для каждой страницы со списком списков
        def dict_page2(li, dic):
            s = {}

            if self.zadacha in ['Объеденить']:
                s['Маршрутная карта'] = self.li   # копирование базовой страницы
            else:
                print('не перезаписываем первую страницу')

            n = 1
            for key in dic.keys():
                if n > int(self.index + 1):
                    nn = 1
                    one = copy.deepcopy(li)
                    for j in dic[key]:
                        if j in ['x', 'X', 'х', 'Х']:
                            one[nn][-1] = 0
                        else:
                            one[nn][-1] = '''='Маршрутная карта'!''' + str(self.alfavit[self.index]) + str(nn + 1) + \
                                              '''-'Маршрутная карта'!''' + str(self.alfavit[n - 1]) + \
                                              str(nn + 1)
                        nn += 1
                    s[key] = one
                n += 1

            return s

        s1 = one_list()     # получаем попловину списка списком
        s2 = two_dict()     # получаем словарь из таблицы
        s3 = dict_page2(s1, s2)
        return s3

class writeToExl:
    def __init__(self, dic, InDate, fname, zadacha):
        self.Nzakaz = InDate[0]         # номер заказа
        self.proj = InDate[1]           # Номер проекта
        self.NameProj = InDate[2]       # Наименование проекта
        self.etap = InDate[3]           # номер этапа
        self.name_proizv = InDate[4]    # наименование произ. полщадки
        self.dic = dic                  # словарь со страницами
        self.fname = fname              # путь к файлу
        self.zadacha = zadacha

    def isInt(self, value):
        """проверяем является ли значение числом"""
        try:
            int(value)
            return True
        except ValueError:
            return False

    def Deleting_sheets(self):
        print(self.fname)
        wb = openpyxl.load_workbook(self.fname)

        for sheet in wb.sheetnames:
            print(sheet)
            if sheet == 'Маршрутная карта':
                print('не удаляем лист')
            else:
                wb.remove(wb[sheet])
                print('удаляем лист')

        wb.save('temporary file.xlsx')
        return 'temporary file.xlsx'

    def Write(self):
        print('ЗАПИСЫВАЕМ ДАННЫЕ В ФАЙЛ')

        if self.zadacha in ['Объеденить']:
            wb = openpyxl.Workbook()   # Создаем виртуальную книгу
        else:
            wb = openpyxl.load_workbook(self.fname)     # презаписываем начальный файл
        for i in self.dic:
            sheet = wb.create_sheet(i)
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToHeight = False
            # верхний левый колонтитул
            sheet.oddHeader.left.text = 'Проект № ' + str(self.proj) + ' ' + str(self.NameProj) + \
                                        '\nЗаказ № ' + str(self.Nzakaz) + \
                                        '\nЭтап № ' + str(self.etap)
            sheet.oddHeader.left.size = 12
            sheet.oddHeader.left.font = "Century Gothic"
            # верхний центральный колонтитул
            sheet.oddHeader.center.text = i
            sheet.oddHeader.center.size = 20
            sheet.oddHeader.center.font = "Century Gothic"
            # верхний правый колонтитул
            now = datetime.datetime.today().strftime("%d %m %Y")
            if self.name_proizv in ['Solaris']:
                sheet.oddHeader.right.text = 'ООО «ПИК «СОЛЯРИС»\n' + str(now)
            else:
                sheet.oddHeader.right.text = 'ООО «Питон Кама»\n' + str(now)
            sheet.oddHeader.right.size = 12
            sheet.oddHeader.right.font = "Century Gothic"
            # нижний левый колонтитул
            sheet.oddFooter.left.text = 'Время начала:____:____' + '\n' + 'Время завершения:____:____'
            sheet.oddFooter.left.size = 14
            sheet.oddFooter.left.font = "Century Gothic"
            # нижний центральный колонтитул
            sheet.oddFooter.center.text = 'ФИО исполнителя:____________'
            sheet.oddFooter.center.size = 14
            sheet.oddFooter.center.font = "Century Gothic"
            y = 1

            for j in self.dic[i]:
                x = 1
                for k in j:
                    cell = sheet.cell(row=y, column=x)
                    cell.font = Font(name='Century Gothic', color='000000')
                    if self.isInt(k) == True:
                        cell.value = int(k)
                    else:
                        cell.value = k
                    x += 1
                y += 1

        # wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        name_out_file = 'Заказ №' + str(self.Nzakaz) + ' Этап №' + str(self.etap) + '.xlsx'
        print('Сохранение в файл: ', name_out_file)
        wb.save(name_out_file)
        return name_out_file



if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    # item = exl_to_list(26, 'all', 'D:/PycharmProjects/cex_vedomost/Маршрутная карта.xlsx', 'Маршрутная карта').reading()
    # item = list_to_dict(item).transformation()
    # writeToExl(item, ['10', 'Первый', '1', '1', 'Солярис']).Write()
    writeToExl([1], [1, 1, 1, 1, 1], 'D:\Загрузки\Заказ №1 Этап №1.xlsx', 1).Deleting_sheets()