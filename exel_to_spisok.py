import copy
import openpyxl

class InOut:
    """модуль в котором мы обрабатывам входной файл из экселя для выдачи обработанных списков для начальников смен
    и для создания контрольной ведомости"""

    def __init__(self, fname, fdat):
        self.fname = fname  # имя файла который нужно открыть
        self.fdat = fdat    # форма которую нужно выдать, для контрольной ведомости 'k', для нач.смен 'n'


    def zapusk(self):

        if self.fdat == 'n':
            print('Ведомость нач. смен')
            return self.slog_povt(self.sklad_del(self.spisok_etapov(self.addition())))
        elif self.fdat == 'k':
            print('Контрольная ведомость')
            return self.slovar_raskritii(self.addition())
        elif self.fdat == 'o':
            print('Отметка нестандарта')
            return self.sortNest()
        else:
            return 'ошибка'

    def addition(self):
        """модуль создания из файла эксель раскрытого списка"""
        print('ЭКСЕЛЬ РАСКРЫТЫЙ СПИСКОК')

        array = []
        wb = openpyxl.load_workbook(self.fname, data_only=True)
        sheet = wb.get_sheet_by_name('Спецификация')
        rows = sheet.max_row
        cols = sheet.max_column

        for i in range(1, rows + 1):
            sel = []

            for j in range(1, cols + 1):
                if j <= 14:
                    cell = sheet.cell(row=i, column=j)
                    sel.append(str(cell.value))
                else:
                    continue

            array.append(sel)
        print(array)

        return array

    def spisok_etapov(self, spis):
        """модуль создания из списка словарь этапов"""
        print('СЛОВАРЬ ЭТАПОВ')
        array = []

        # перебираем строк и создаем полный список
        s = 0
        del spis[0]  # удаляем первую строку

        for i in spis:
            if i[0] in ('Итог', 'Итог:', 'Итого', 'Итого:', 'итог', 'итог:', 'итого', 'итого:', 'None'):
                print('удаляем строчку ', spis.index(i))
                del spis[spis.index(i)]
                continue
            else:
                print(i)
                kol_povt = int(i[11])
                i.pop(11)

            while kol_povt >= 1:
                array.append(i)
                kol_povt = kol_povt - 1

        # меняем формат значения этапа на цифренное
        for s in array:
            if s[12] in ('None', '', ' '):
                s[12] = 1
            else:
                s[12] = int(s[12])
        print(array)
        # создаем словарь этапов из полного списка
        array = sorted(array, key=lambda ji: ji[12])  # сортируем итоговый список по номеру этапа
        s = int(array[0][12])  # номер этапа
        di = {s: [array[0]]}  # создаем словарь с этапами
        del array[0]  # удаляем первую строчку, которую добавили в словарь

        for i in array:

            if int(i[12]) == s:
                di[s].append(i)
            else:
                s += 1
                di[s] = [i]

        return di

    def sklad_del(self, di):
        """модуль убирает все строчки связаннные со складом"""
        print('УДАЛЯЕМ СКЛАД')

        for i in di:
            s = []

            for j in di[i]:

                if j[11] != 'склад':
                    s.append(j)

            di[i] = s


        return di

    def slog_povt(self, di):
        """модуль складывает все повторяющиеся элементы списка"""
        print('СКЛАДЫВАЕМ ПОВТОРЫ')

        def del_n(di):
            # убираем ненужные атрибуты
            n = 0
            while len(di) > n:
                dj = copy.copy(di[n])
                del dj[0]
                s = 1

                while s <= 3:
                    del dj[9]
                    s += 1

                dj.append(1)
                di[n] = dj
                n += 1

            return di

        for i in di:
            di[i] = del_n(di[i])
            # начинаем сравнивать
            kol_str = len(di[i])  # записываем клличесвто элементов в списке
            array = []

            if kol_str == 1:
                print('no sum')
                continue
            else:
                for j in di[i]:
                    q = copy.copy(j)
                    #print('список который сравниваем', q)
                    qN = q[-1]  # присваиваем значение количества строки которую нужно сравнить
                    del q[-1]   # удаляем значение количества строки которую нужно сравнить
                    #print('Теперь у нас есть список', q)
                    y = di[i].index(j) + 1  # нам нужна следующая строка

                    # начинаем проверять на совпадение
                    while y <= kol_str-1:   # начинаем сравнивать со следующей строкой
                        ij = 0
                        w = copy.copy(di[i][y])     # строка с которой сравниваем это следующая строка
                        #print('список с которым сравниваем', w)
                        wN = w[-1]  # присваиваем значение количества строки с которой нужно сравнить
                        del w[-1]   # удаляем значение количества строки с которой нужно сравнить
                        #print('Теперь у нас есть список', w)

                        if q == w:
                            ij += 1
                            #print(qN)
                            qN = int(wN) + int(qN)  # складываем количество
                            #print('Складываем количество +', wN, ' Итог=', qN)
                            del di[i][y]
                            kol_str = len(di[i])
                        else:
                            #print('элемент не похож')
                            y = y + 1

                    q.insert(9, qN)
                    array.append(q)
                    #print(array)

                di[i].clear()
                di[i] = array
        return di

#
    def slovar_raskritii(self, array):
        """модуль для создания словаря из раскрытого списка"""
        print('РАСКРЫТЫЙ СЛОВАРЬ')

        def si(spis):
            # перебираем строк и создаем полный список
            array = []
            s = 0

            for i in spis:

                if s == 0:
                    i.pop(11)
                    s += 1
                    kol_povt = 1
                elif i[7] in ['kz', 'op', 'ks', 'pp', 'gp', 'ad', 'ksb']:
                    i.append(i[11])
                    i.pop(11)
                    s += 1
                    kol_povt = 1
                elif i[0] in ('Итог', 'Итог:', 'Итого', 'Итого:', 'итог', 'итог:', 'итого', 'итого:', 'None'):
                    print('удаляем сточку ',  spis.index(i))
                    del spis[spis.index(i)]
                    continue
                else:
                    print(i)
                    kol_povt = int(i[11])
                    i.pop(11)

                while kol_povt >= 1:
                    array.append(i)
                    kol_povt = kol_povt - 1
            return array

        n = 0
        s = {}
        array = si(array)

        for i in array:

            if i[12] in ('None', '', ' ', None):
                i[12] = 1
            if n == 0:
                n += 1
                continue
            else:
                nn = 0
                s[n] = {}

                if i[7] in ['kz', 'op', 'ks', 'pp', 'gp', 'ad', 'ksb']:
                    s[n]['Кол, шт'] = i[13]
                    del i[13]

                for j in i:

                    if j == 'None':
                        j = ''
                    s[n][array[0][nn]] = j
                    nn += 1

            n += 1

        return s

    def sortNest (self):
        """модуль для отметки нестандарта в контрольной ведомости"""
        print('СОРТРОВКА НЕСТАНДАРТА')
        s1 = self.fname[0]
        s2 = self.fname[1]
        print('s1', s1)
        print('s2', s2)

        for i in s1:
            f = [s1[i]['Серия'], s1[i]['ip'], s1[i]['Материал'], s1[i]['Кол. пров.'], s1[i]['In'],
                 s1[i]['Наименование'], s1[i]['Обозначение'], s1[i]['тип'], s1[i]['размер']]

            for i2 in s2:

                if s2[i2]['Ном. ток, А'] in ['', '-', '0', 0, ' ']:
                    s2[i2]['Ном. ток, А'] = 0
                print(s2[i2]['Этап'])
                if s2[i2]['Этап'] in ['None', '', ' ']:
                    s2[i2]['Этап'] = 1

                f2 = [s2[i2]['Серия'], s2[i2]['IP'], s2[i2]['Мат. Пров.'], s2[i2]['Кол. Пров.'],
                      s2[i2]['Ном. ток, А'], s2[i2]['Наименование'], s2[i2]['Обозначение'],
                      s2[i2]['Тип'], s2[i2]['Размер, мм']]

                if f == f2:
                    s2[i2]['Примечание'] = 'СИ'
                    s2[i2]['количество'] = s1[i]['количество']

        return s2

if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    print(InOut('пример.xlsx', 'k').zapusk())
